#!/usr/bin/env python3
"""東京都中央卸売市場 月報明細データ クロール

Usage:
  python crawl_wholesale.py --backfill           # 202502〜直近 全月取得（初回）
  python crawl_wholesale.py                       # 未取得月のみ追加（月次運用）
  python crawl_wholesale.py --month 202604       # 特定月のみ
  python crawl_wholesale.py --file path.xlsx     # ローカルxlsxを直接読込
  python crawl_wholesale.py --force              # 既存月も上書き

出力: docs/fish-value/wholesale-prices.json
"""
from __future__ import annotations
import argparse
import io
import json
import os
import re
import sys
import time
import urllib.request
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
SPECIES_MAP = HERE / 'fish-species-map.json'
OUTPUT = HERE / 'wholesale-prices.json'

GEPPO_PAGE = 'https://www.shijou.metro.tokyo.lg.jp/torihiki/geppo'
GEPPO_HOST = 'https://www.shijou.metro.tokyo.lg.jp'
UA = 'Mozilla/5.0 (compatible; funatsuri-yoso-bot/1.0; +https://funatsuri-yoso.com)'
WAIT_SEC = 0.8

TARGET_CATEGORIES = {'鮮　魚', '活魚類'}  # 鮮魚 + 活魚のみ（冷凍/加工は除外）


def fetch(url: str, binary: bool = False):
    req = urllib.request.Request(url, headers={'User-Agent': UA, 'Accept': '*/*'})
    with urllib.request.urlopen(req, timeout=120) as r:
        data = r.read()
        return data if binary else data.decode('utf-8', errors='ignore')


def list_available_months() -> dict[str, str]:
    """月報ページから s{YYYYMM}meisai_{YYYYMMDD}.xlsx の絶対URLを抽出"""
    html = fetch(GEPPO_PAGE)
    pat = re.compile(r'(?:href=["\'])?([^"\'\s<>]*s(\d{6})meisai_(\d{8})\.xlsx)')
    found: dict[str, str] = {}
    for m in pat.finditer(html):
        href, yyyymm, _ = m.group(1), m.group(2), m.group(3)
        if href.startswith('//'):
            href = 'https:' + href
        elif href.startswith('/'):
            href = GEPPO_HOST + href
        elif not href.startswith('http'):
            href = GEPPO_PAGE.rstrip('/') + '/' + href.lstrip('/')
        found.setdefault(yyyymm, href)
    return found


def parse_xlsx(xlsx_bytes: bytes) -> tuple[str, dict]:
    """xlsxバイトから (yyyymm, 品目集計dict) を返す"""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    yyyymm: str | None = None
    by_item: dict[str, dict] = defaultdict(lambda: {'qty': 0, 'amount': 0, 'unit_prices': []})

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[5] is None:
            continue
        if yyyymm is None and row[1] is not None:
            yyyymm = str(int(row[1]))
        if row[7] not in TARGET_CATEGORIES:
            continue
        item = row[11]
        qty = row[14] or 0
        amt = row[15] or 0
        if qty <= 0 or item is None:
            continue
        by_item[item]['qty'] += qty
        by_item[item]['amount'] += amt
        by_item[item]['unit_prices'].append(amt / qty)
    if yyyymm is None:
        raise ValueError('yyyymm 列を取得できず')
    return yyyymm, by_item


def load_item_to_pfid() -> dict[str, str]:
    """fish-species-map.json から geppo_item -> price_fish_id の逆引き辞書"""
    with open(SPECIES_MAP, encoding='utf-8') as f:
        m = json.load(f)
    item_to_pfid: dict[str, str] = {}
    for s in m['species']:
        gi = s.get('geppo_item')
        if gi:
            item_to_pfid.setdefault(gi, s['price_fish_id'])
    return item_to_pfid


def aggregate(yyyymm: str, by_item: dict, item_to_pfid: dict[str, str]) -> dict:
    prices = []
    for item, data in by_item.items():
        unit = data['unit_prices']
        if not unit:
            continue
        prices.append({
            'price_fish_id': item_to_pfid.get(item),  # null許容（マッピング外も保持）
            'geppo_item': item,
            'qty_kg': int(data['qty']),
            'amount_yen': int(data['amount']),
            'avg_yen_per_kg': round(data['amount'] / data['qty']),
            'high_yen_per_kg': round(max(unit)),
            'low_yen_per_kg': round(min(unit)),
            'shipper_count': len(unit),
        })
    prices.sort(key=lambda x: (x['price_fish_id'] or 'zzz', x['geppo_item']))
    return {
        'yyyymm': yyyymm,
        'fetched_at': datetime.now().astimezone().isoformat(timespec='seconds'),
        'prices': prices,
    }


def load_existing() -> dict:
    if OUTPUT.exists():
        with open(OUTPUT, encoding='utf-8') as f:
            return json.load(f)
    return {
        'version': 'v1',
        'source': '東京都中央卸売市場 月報 明細データ',
        'source_url': GEPPO_PAGE,
        'updated_at': None,
        'months': [],
    }


def save(doc: dict) -> None:
    doc['updated_at'] = date.today().isoformat()
    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(doc, f, ensure_ascii=False, indent=2)


def upsert_month(doc: dict, month_data: dict) -> None:
    doc['months'] = [m for m in doc['months'] if m['yyyymm'] != month_data['yyyymm']]
    doc['months'].append(month_data)
    doc['months'].sort(key=lambda x: x['yyyymm'])


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument('--backfill', action='store_true', help='202502〜直近 全月取得')
    p.add_argument('--month', help='特定月 yyyymm のみ取得')
    p.add_argument('--file', help='ローカル xlsx を直接読込（URL fetch しない）')
    p.add_argument('--force', action='store_true', help='既存月も上書き')
    args = p.parse_args()

    item_to_pfid = load_item_to_pfid()
    doc = load_existing()
    existing = {m['yyyymm'] for m in doc['months']}

    if args.file:
        with open(args.file, 'rb') as f:
            xlsx_bytes = f.read()
        yyyymm, by_item = parse_xlsx(xlsx_bytes)
        if not args.force and yyyymm in existing:
            print(f'[{yyyymm}] 既存 → skip (--force で上書き)')
            return 0
        md = aggregate(yyyymm, by_item, item_to_pfid)
        md['source_file'] = os.path.basename(args.file)
        upsert_month(doc, md)
        save(doc)
        print(f'[{yyyymm}] 品目 {len(md["prices"])} 件 → {OUTPUT.name}')
        return 0

    available = list_available_months()
    print(f'月報ページからリンク検出: {len(available)} 月分')
    if not available:
        print('リンク取得失敗（ページ構造変更の可能性）', file=sys.stderr)
        return 1

    if args.month:
        targets = [args.month] if args.month in available else []
        if not targets:
            print(f'{args.month} はページに見つからず', file=sys.stderr)
            return 1
    elif args.backfill:
        targets = sorted(available)
    else:
        targets = sorted(set(available) - existing)

    if not args.force:
        targets = [m for m in targets if m not in existing]

    if not targets:
        print('取得対象なし（最新まで取得済）')
        return 0

    print(f'取得対象: {targets}')
    for ym in targets:
        url = available[ym]
        print(f'[{ym}] DL: {url}')
        try:
            xlsx_bytes = fetch(url, binary=True)
        except Exception as e:
            print(f'  失敗: {e}', file=sys.stderr)
            continue
        time.sleep(WAIT_SEC)

        yyyymm, by_item = parse_xlsx(xlsx_bytes)
        md = aggregate(yyyymm, by_item, item_to_pfid)
        md['source_url'] = url
        md['source_file'] = url.rsplit('/', 1)[-1]
        upsert_month(doc, md)
        save(doc)
        print(f'  品目 {len(md["prices"])} 件')

    print(f'\n総月数: {len(doc["months"])} → {OUTPUT}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
